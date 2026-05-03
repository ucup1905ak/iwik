import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    DEFAULT_FONT
)


def export_to_xlsx(products: list, filename: str = None) -> tuple[bool, str]:
    """
    Export products to Excel file (.xlsx)
    
    Args:
        products: List of Product namedtuples or dicts with product data
        filename: Output Excel filename (defaults to 'products_export_TIMESTAMP.xlsx')
    
    Returns:
        Tuple of (success: bool, message: str)
    """
    try:
        # Create assets/xlsx directory if it doesn't exist
        base_path = os.path.dirname(os.path.dirname(__file__))
        xlsx_dir = os.path.join(base_path, 'assets', 'xlsx')
        os.makedirs(xlsx_dir, exist_ok=True)
        
        if filename is None:
            timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            filename = f'products_export_{timestamp}.xlsx'
        
        filepath = os.path.join(xlsx_dir, filename)
        
        if not products:
            return False, "Data kosong, tidak ada yang diekspor"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Produk"
        
        # Define styles
        header_fill = PatternFill(start_color="4F6EF7", end_color="4F6EF7", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3")
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Headers
        headers = ["No", "Nama Produk", "Merek", "SKU", "Kategori", "Harga (Rp)", "Stok"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 10
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Data rows
        for idx, product in enumerate(products, 2):
            # Handle both namedtuple and dict
            if hasattr(product, '_asdict'):
                data = product._asdict()
            else:
                data = product
            
            row_data = [
                idx - 1,  # No
                data.get('name', ''),
                data.get('brand', ''),
                data.get('sku', ''),
                data.get('category', ''),
                data.get('price', 0),
                data.get('stock', 0),
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col_num)
                cell.value = value
                cell.border = border
                
                # Format number columns
                if col_num in [6]:  # Price
                    cell.number_format = '#,##0'
                    cell.alignment = center_alignment
                elif col_num in [1, 7]:  # No, Stock
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
        
        # Save
        wb.save(filepath)
        return True, f"Excel berhasil dibuat: {filepath}"
        
    except Exception as e:
        return False, f"Error exporting to Excel: {str(e)}"


def import_from_xlsx(filepath: str) -> tuple[bool, list, str]:
    """
    Import products from Excel file (.xlsx)
    
    Args:
        filepath: Path to Excel file
    
    Returns:
        Tuple of (success: bool, products: list[dict], message: str)
        products contains dicts with keys: name, brand, sku, category, price, stock
    """
    try:
        if not os.path.exists(filepath):
            return False, [], f"File tidak ditemukan: {filepath}"
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        if not ws:
            return False, [], "Workbook tidak memiliki sheet aktif"
        
        products = []
        errors = []
        
        # Skip header (row 1)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue  # Skip empty rows
            
            try:
                # Expected columns: No, Name, Brand, SKU, Category, Price, Stock
                _, name, brand, sku, category, price, stock = row[:7]
                
                # Validation
                if not name or str(name).strip() == "":
                    errors.append(f"Baris {row_num}: Nama produk kosong")
                    continue
                
                if not sku or str(sku).strip() == "":
                    errors.append(f"Baris {row_num}: SKU kosong")
                    continue
                
                # Type conversion
                try:
                    price = float(price) if price else 0
                    stock = int(stock) if stock else 0
                except (ValueError, TypeError):
                    errors.append(f"Baris {row_num}: Harga atau stok tidak valid")
                    continue
                
                product = {
                    'name': str(name).strip(),
                    'brand': str(brand).strip() if brand else "",
                    'sku': str(sku).strip(),
                    'category': str(category).strip() if category else "",
                    'price': price,
                    'stock': stock,
                }
                
                products.append(product)
                
            except (ValueError, TypeError, IndexError) as e:
                errors.append(f"Baris {row_num}: Format data tidak valid - {str(e)}")
                continue
        
        if not products:
            error_msg = "\n".join(errors) if errors else "Tidak ada data produk ditemukan"
            return False, [], error_msg
        
        success_msg = f"Berhasil membaca {len(products)} produk"
        if errors:
            success_msg += f"\n\nPeringatan:\n" + "\n".join(errors[:5])  # Show first 5 errors
            if len(errors) > 5:
                success_msg += f"\n... dan {len(errors) - 5} error lainnya"
        
        return True, products, success_msg
        
    except Exception as e:
        return False, [], f"Error importing from Excel: {str(e)}"
